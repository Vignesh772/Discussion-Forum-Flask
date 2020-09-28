# Note we imported request!
from flask import Flask
from flask import render_template
from flask import request
import pandas as pd
import numpy as np
import pickle
import time
import openpyxl

from datetime import datetime
#from scipy import sparse
# Load vectorizer and similarity measure
#from sklearn.feature_extraction import text
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

df=pd.read_excel('disc_dataset.xlsx')
recent_question=df.tail(5)
recent_question=recent_question.iloc[:,1].values

posted_question=''
suggestion=[]
suggestion_answer=[]


wb = openpyxl.load_workbook('disc_dataset.xlsx')
similar_id=[]

sheet= wb.get_active_sheet()
recent_id=[i for i in range((sheet.max_row-5),sheet.max_row)]


app = Flask(__name__)

@app.route('/')
def home():
    return render_template('home.html')



@app.route('/signup')
def signup():
    return render_template('signup_page.html')



@app.route('/generalforum')
def generalforum():

    #print(recent_question)
    similar_questions=None
    inputQuestion=None
    submit_status=request.args.get('submit_status')
    if(submit_status=='Submit'):
        inputQuestion = request.args.get('inputQuestion')
        global posted_question
        posted_question=inputQuestion
        similar_questions=model(inputQuestion)



    return render_template('forum.html',submit_status=submit_status,similar_questions=similar_questions,
                            inputQuestion=inputQuestion,recent_question=recent_question,similar_id=similar_id)



@app.route('/threadpage',methods=['GET','POST'])
def thread_page():
    id = int(request.args.get('values'))
    type=(request.args.get('type'))
    if(type=='similar'):
        return render_template('thread.html',recent_question=recent_question,suggestion=suggestion[similar_id.index(id)],answer=suggestion_answer[similar_id.index(id)])
    elif(type=='recent'):
        return render_template('thread.html',recent_question=recent_question,suggestion=recent_question[id],answer=df.Answer.iloc[[recent_id[id]]].values[0])






@app.route('/submitthread')
def submit_thread_page():
    global posted_question
    global sheet
    sheet['B'+str(sheet.max_row+1)]=posted_question
    sheet['A'+str(sheet.max_row)]=(sheet.max_row)
    wb.save('disc_dataset.xlsx')

    return render_template('submit_thread.html',recent_question=recent_question,posted_question=posted_question)



def model(ques):
	#print(t1)
    vectorizer1=pickle.load(open("model2.pk", 'rb'))
    Question_vectors=pickle.load(open("question_vector2.pk", 'rb'))

    input_question = ques

	# Locate the closest question
    input_question_vector = vectorizer1.transform([input_question])

	# Compute similarities
    similarities = cosine_similarity(input_question_vector, Question_vectors)
    l=(similarities.shape[1])
    mid=l//2
    sim=similarities[0].argsort()[-5:][::-1]
    for s in sim:
        suggestion.append(df.Question.iloc[[s]].values[0])
        suggestion_answer.append(df.Answer.iloc[[s]].values[0])
        similar_id.append(s)

	    #return("Instructor's reply: " + df.Answer.iloc[closest].values[0])



    return(suggestion)
# This page will be the page after the form



if __name__ == '__main__':
    app.run(debug=True)
